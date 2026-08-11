"""
Execution MCP 관리 API - 등록 커맨드(execution_commands) + 실행 감사로그.
구 commands.py(커맨드 카탈로그)와 system.py(화이트리스트)를 하나로 합친 것이다(#111).

등록 커맨드의 인자 설계:
  관리자는 `head -n {lines} {path}`처럼 **자리표시자가 든 커맨드 한 줄**만 적는다. 콘솔이
  자리표시자를 뽑아 인자 표를 만들고, 각 인자의 타입/필수/기본값/설명을 채우게 한다.
  argv JSON을 손으로 쓰게 하던 예전 방식보다 훨씬 쉽고, 카탈로그의 `{user_id}` 문법과도 같다.
  `{user_id}`는 예약어라 표에 나오지 않는다(호출자 신원에서 자동 주입).

**커맨드는 전부 여기 등록분 하나다**(#128). 예전에는 파이썬 함수로 박아 둔 '내장 커맨드' 7개가
따로 있어서 편집도 삭제도 안 됐는데, 그 7개는 전부 LLM이 이미 아는 표준 리눅스 명령이라
run_command로 실행하면 그만이다. 목록에서 없앴더니 매 요청 프롬프트도 그만큼 가벼워졌다.
"""
import io
import json
import re
import sys
import os

from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import Response
from pydantic import BaseModel

from auth import require_admin
from config_store import get_config
from db import get_pool
from cleaning import clean_text, clean_options_from_dict
from spreadsheet import TABLE_EXTS, read_table_meta, load_table_rows
from uploads import (
    create_upload_session, get_upload_session, delete_upload_session, load_options,
)
from server_files import read_upload_or_server_file

sys.path.append(os.path.join(os.path.dirname(__file__), "../../../shared"))
from execution_exec import (  # noqa: E402
    DEFAULT_DENY_CSV, deny_set, placeholders_in, tool_name_for, validate_definition,
)

router = APIRouter(prefix="/api/execution", tags=["execution"])

_DSN = "execution_db_dsn"
# 툴 이름이 겹치면 안 되는 예약어. run_command는 MCP가 항상 노출하는 미등록 커맨드 실행 툴이다.
_RESERVED_TOOL_NAMES = {"run_command"}


async def _deny() -> set:
    return deny_set(await get_config("execution_deny_commands", DEFAULT_DENY_CSV))


def _row(r) -> dict:
    d = dict(r)
    if isinstance(d.get("args"), str):
        d["args"] = json.loads(d["args"])
    return d


# ---------------------------------------------------------------- 등록 커맨드
class ArgIn(BaseModel):
    name: str
    type: str = "str"                 # str | int | enum
    required: bool = False
    default: str = ""
    description: str = ""
    choices: list[str] = []


class CommandIn(BaseModel):
    """title: 사람이 읽는 이름(한글 가능). tool_name은 서버가 만들어 준다(ASCII 규칙).
    exec_command: `myquota` / `head -n {lines} {path}` 형태의 커맨드 한 줄."""
    title: str
    description: str = ""
    exec_command: str
    args: list[ArgIn] = []
    host_mode: str = "login_server"
    enabled: bool = True
    required_roles: list[str] = []


@router.get("/commands")
async def list_commands(admin: str = Depends(require_admin)):
    pool = await get_pool(_DSN)
    rows = await pool.fetch(
        "SELECT id, tool_name, title, description, exec_command, args, "
        "host_mode, enabled, required_roles, updated_by, updated_at "
        "FROM execution_commands ORDER BY title")
    return [_row(r) for r in rows]


@router.post("/commands/parse")
async def parse_command(body: dict, admin: str = Depends(require_admin)):
    """커맨드 한 줄에서 자리표시자를 뽑아 준다. 콘솔이 입력 중에 인자 표를 만드는 데 쓴다.
    `{user_id}`는 시스템이 자동 주입하므로 표에 넣지 않는다."""
    names = [n for n in placeholders_in(body.get("exec_command") or "") if n != "user_id"]
    return {"placeholders": names,
            "has_user_id": "user_id" in placeholders_in(body.get("exec_command") or "")}


async def _validate(body: CommandIn, existing: set[str], tool_name: str):
    args = [a.model_dump() for a in body.args]
    try:
        validate_definition(tool_name, body.exec_command, args, body.host_mode,
                            await _deny(), existing)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return args


@router.post("/commands")
async def create_command(body: CommandIn, admin: str = Depends(require_admin)):
    pool = await get_pool(_DSN)
    taken = {r["tool_name"] for r in await pool.fetch("SELECT tool_name FROM execution_commands")}
    taken |= _RESERVED_TOOL_NAMES
    tool_name = tool_name_for(body.title, taken, body.exec_command)
    args = await _validate(body, taken, tool_name)
    try:
        row_id = await pool.fetchval(
            """
            INSERT INTO execution_commands
                (tool_name, title, description, exec_command, args,
                 host_mode, enabled, required_roles, updated_by)
            VALUES ($1,$2,$3,$4,$5::jsonb,$6,$7,$8,$9) RETURNING id
            """,
            tool_name, body.title.strip(), body.description.strip(), body.exec_command.strip(),
            json.dumps(args, ensure_ascii=False), body.host_mode,
            body.enabled, [r.strip() for r in body.required_roles if r.strip()], admin)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(400, f"등록 실패 (이름 중복 가능): {e}")
    return {"id": row_id, "tool_name": tool_name, "restart_required": True}


@router.patch("/commands/{command_id}")
async def update_command(command_id: int, body: CommandIn, admin: str = Depends(require_admin)):
    pool = await get_pool(_DSN)
    cur = await pool.fetchrow(
        "SELECT tool_name FROM execution_commands WHERE id = $1", command_id)
    if not cur:
        raise HTTPException(404, "커맨드를 찾을 수 없습니다.")
    taken = {r["tool_name"] for r in await pool.fetch(
        "SELECT tool_name FROM execution_commands WHERE id <> $1", command_id)}
    args = await _validate(body, taken | _RESERVED_TOOL_NAMES, cur["tool_name"])
    await pool.execute(
        """
        UPDATE execution_commands
        SET title=$2, description=$3, exec_command=$4, args=$5::jsonb,
            host_mode=$6, enabled=$7, required_roles=$8, updated_by=$9, updated_at=now()
        WHERE id=$1
        """,
        command_id, body.title.strip(), body.description.strip(), body.exec_command.strip(),
        json.dumps(args, ensure_ascii=False), body.host_mode,
        body.enabled, [r.strip() for r in body.required_roles if r.strip()], admin)
    # enabled/역할은 실행 시점에 읽으므로 즉시 반영된다. 나머지는 툴 스키마라 재시작이 필요하다.
    return {"ok": True, "restart_required": True}


@router.patch("/commands/{command_id}/enabled")
async def toggle_command(command_id: int, body: dict, admin: str = Depends(require_admin)):
    """활성/비활성만 바꾼다. **실시간 반영**이라 재시작이 필요 없다."""
    pool = await get_pool(_DSN)
    row = await pool.fetchrow(
        "UPDATE execution_commands SET enabled=$2, updated_by=$3, updated_at=now() "
        "WHERE id=$1 RETURNING id", command_id, bool(body.get("enabled")), admin)
    if not row:
        raise HTTPException(404, "커맨드를 찾을 수 없습니다.")
    return {"ok": True, "restart_required": False}


@router.delete("/commands/{command_id}")
async def delete_command(command_id: int, admin: str = Depends(require_admin)):
    pool = await get_pool(_DSN)
    await pool.execute("DELETE FROM execution_commands WHERE id = $1", command_id)
    return {"ok": True, "restart_required": True}


# ---------------------------------------------------------------- 실행 로그
@router.get("/logs")
async def list_logs(limit: int = 100, admin: str = Depends(require_admin)):
    pool = await get_pool(_DSN)
    rows = await pool.fetch(
        "SELECT id, tool_name, params, requested_by, status, result, created_at "
        "FROM job_logs ORDER BY created_at DESC LIMIT $1", max(1, min(int(limit), 500)))
    return [dict(r) for r in rows]


# ---------------------------------------------------------------- 엑셀 양식(내려받기)
# 한 커맨드 = 한 행. 인자는 열로 펼친다(`인자1 …` ~ `인자4 …`).
#
# 왜 이 모양인가: 예전 일괄 등록은 이름·설명·실행 커맨드 세 열만 읽고 **인자를 아예 버렸다**.
# 그래서 `phd info {option} {job_id}`처럼 자리표시자가 있는 커맨드는 엑셀로 넣어 봐야 인자가
# 빈 채로 등록됐고, 결국 한 건씩 화면에서 다시 채워야 했다(사용자 지적, #140).
TEMPLATE_ARG_SLOTS = 4
_ARG_FIELDS = ["이름", "설명", "타입", "필수", "기본값", "선택지"]
_BASE_COLUMNS = ["이름", "설명", "실행 커맨드", "실행 위치", "활성", "필요 역할"]

# `인자2 선택지` 같은 머리글을 (번호, 항목)으로 읽는다. 공백은 있어도 없어도 된다.
_ARG_HEADER_RE = re.compile(r"^\s*인자\s*(\d+)\s*(이름|설명|타입|필수|기본값|선택지)\s*$")

# 선택지 구분자: 셀 안 줄바꿈 또는 `|`. 콤마는 쓰지 않는다 - 설명에 콤마가 흔하다
# ("-j: Return job info, json format").
_CHOICE_SPLIT = re.compile(r"[\n|]+")

_TEMPLATE_EXAMPLES = [
    ["myquota", "사용자의 홈 스토리지 디렉토리, 할당 용량, 현재 사용 현황을 반환합니다.",
     "myquota", "로그인 서버", "Y", "", *[""] * (TEMPLATE_ARG_SLOTS * len(_ARG_FIELDS))],
    ["s2_phd_list", "사용자가 S2 스케줄러로 실행한 GPU batch job 목록을 반환합니다.",
     "phd list {option}", "로그인 서버", "Y", "",
     "option", "출력 형식 옵션. 비우면 요약 목록.", "선택형", "N", "",
     "-l: 상세 정보를 길게 출력\n-lf: 선택 가능한 필드 목록 출력",
     *[""] * ((TEMPLATE_ARG_SLOTS - 1) * len(_ARG_FIELDS))],
    ["s2_phd_info", "S2 스케줄러 GPU batch job id의 상세 정보를 반환합니다.",
     "phd info {option} {job_id}", "로그인 서버", "Y", "",
     "option", "출력 형식 옵션.", "선택형", "N", "",
     "-j: JSON 형식으로 반환\n-tl: 부가 정보까지 출력\n-lf: 선택 가능한 필드 목록",
     "job_id", "S2 스케줄러 job id. s2_phd_list로 얻습니다.", "문자열", "Y", "", "",
     *[""] * ((TEMPLATE_ARG_SLOTS - 2) * len(_ARG_FIELDS))],
]

_TEMPLATE_GUIDE = [
    "■ 한 행이 커맨드 하나입니다. '이름'이 같으면 덮어씁니다(수정하려면 이름을 그대로 두세요).",
    "■ 실행 커맨드의 {이름} 자리가 인자가 됩니다. 예: phd info {option} {job_id}",
    "   → {option}이 인자1, {job_id}가 인자2 입니다. '인자N 이름'을 비워 두면 이 순서로 자동 연결됩니다.",
    "■ 타입 / 필수 / 활성 / 실행 위치 칸은 **드롭다운**입니다. 셀을 클릭하면 목록이 나옵니다.",
    "   직접 타이핑하면 거부됩니다(오타로 다른 뜻이 되는 것을 막기 위함).",
    "■ 인자 타입: 문자열 / 정수 / 선택형. 비우면 문자열입니다.",
    "■ '선택형'일 때만 '선택지'를 채웁니다. 한 줄에 하나씩(또는 | 로 구분), '값: 설명' 형태로 씁니다.",
    "   예)  -j: JSON 형식으로 반환      ← '-j'가 실제로 붙는 값, 뒤는 에이전트가 읽는 설명",
    "   콜론 뒤에 **공백**이 있어야 값과 설명이 갈립니다. 콤마로 나누지 마세요.",
    "■ 설명은 에이전트가 이 커맨드를 고르는 유일한 근거입니다. 무엇을 돌려주는지 한두 줄로 쓰세요.",
    "■ 필수: Y = 에이전트가 반드시 채워야 함. N = 비워도 됨(비우면 Y로 취급하지 않습니다).",
    "■ 활성: 비우면 Y. 실행 위치: 비우면 '로그인 서버'.",
    "■ 필요 역할: 비우면 **누구나** 실행할 수 있습니다. 값을 넣으면 그 역할만 실행 가능.",
    "■ 기본값: **비워 두면 그 인자는 커맨드에서 통째로 빠집니다.**",
    "   예) 'phd list {option}' 에서 option 이 비면 실제 실행은 'phd list' 입니다.",
    "   자주 쓰는 값이 있으면 여기 넣으세요(에이전트가 값을 안 주면 이 값이 쓰입니다).",
    "■ {user_id}는 쓰지 않아도 됩니다. 실행 계정은 시스템이 질문한 사용자로 강제 주입합니다.",
    "■ 업로드 후 execution-mcp 재시작이 필요합니다.",
]


def _template_columns() -> list[str]:
    cols = list(_BASE_COLUMNS)
    for i in range(1, TEMPLATE_ARG_SLOTS + 1):
        cols += [f"인자{i} {f}" for f in _ARG_FIELDS]
    return cols


# 드롭다운으로 고를 값. 자유 입력을 막아야 오타로 조용히 다른 뜻이 되는 것을 방지한다
# ("선택"이라고 적으면 선택형이 아니라 문자열로 들어가는 식).
_DROPDOWNS = {
    "타입": ["문자열", "정수", "선택형"],
    "필수": ["Y", "N"],
    "활성": ["Y", "N"],
    "실행 위치": ["로그인 서버", "대상 서버"],
}
# 역할은 값이 환경마다 달라 고정 목록을 만들 수 없다. 흔한 것만 제안하고 직접 입력도 허용한다.
_ROLE_SUGGESTIONS = ["", "admin", "user"]

# 엑셀 목록 검증은 최대 255자다. 넘으면 시트에 숨김 열을 두고 참조해야 하는데,
# 우리 목록은 전부 짧아서 인라인으로 충분하다.
_MAX_INLINE_VALIDATION = 255


def _validation(values: list[str], allow_free: bool = False):
    from openpyxl.worksheet.datavalidation import DataValidation

    formula = '"' + ",".join(values) + '"'
    assert len(formula) <= _MAX_INLINE_VALIDATION, formula
    dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                        showDropDown=False)      # False = 드롭다운 화살표를 **보여준다**
    if not allow_free:
        dv.errorStyle = "stop"
        dv.error = "목록에 있는 값 중에서 고르세요."
        dv.errorTitle = "허용되지 않는 값"
    else:
        dv.errorStyle = "warning"
    dv.prompt = "목록에서 고르세요: " + ", ".join(v or "(비움)" for v in values)
    dv.promptTitle = "선택"
    dv.showInputMessage = True
    dv.showErrorMessage = True
    return dv


def _build_workbook(rows: list[list]) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "커맨드"
    columns = _template_columns()
    ws.append(columns)
    head_fill = PatternFill("solid", fgColor="DDE6F0")
    for idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        # 기본 열은 넓게, 인자 열은 항목에 맞춰. 설명·선택지는 줄바꿈이 들어간다.
        wide = name in ("설명", "실행 커맨드") or name.endswith(("설명", "선택지"))
        ws.column_dimensions[cell.column_letter].width = 42 if wide else 14
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    # 드롭다운(데이터 유효성). 머리글 이름으로 열을 찾아 붙인다 - `인자3 타입`처럼
    # 번호가 붙은 열도 같은 규칙으로 잡힌다.
    last_row = max(len(rows) + 1, 400)          # 빈 양식에도 넉넉히 걸어 둔다
    for idx, name in enumerate(columns, start=1):
        field = name.split(" ", 1)[-1] if name.startswith("인자") else name
        values = _DROPDOWNS.get(field)
        allow_free = False
        if values is None and field == "필요 역할":
            values, allow_free = _ROLE_SUGGESTIONS, True
        if values is None:
            continue
        dv = _validation(values, allow_free)
        ws.add_data_validation(dv)
        col = get_column_letter(idx)
        dv.add(f"{col}2:{col}{last_row}")

    guide = wb.create_sheet("작성 방법")
    guide.column_dimensions["A"].width = 110
    for line in _TEMPLATE_GUIDE:
        guide.append([line])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _arg_type_label(kind: str) -> str:
    return {"str": "문자열", "int": "정수", "enum": "선택형"}.get(kind, "문자열")


def _command_to_row(c: dict) -> list:
    row = [c.get("title") or "", c.get("description") or "", c.get("exec_command") or "",
           "로그인 서버" if (c.get("host_mode") or "login_server") == "login_server" else "대상 서버",
           "Y" if c.get("enabled") else "N",
           ", ".join(c.get("required_roles") or [])]
    args = c.get("args") or []
    for i in range(TEMPLATE_ARG_SLOTS):
        a = args[i] if i < len(args) else {}
        row += [a.get("name") or "", a.get("description") or "",
                _arg_type_label(a.get("type") or "str"),
                "Y" if a.get("required") else "N", a.get("default") or "",
                "\n".join(str(c2) for c2 in (a.get("choices") or []))]
    return row


@router.get("/commands/template.xlsx")
async def download_template(mode: str = "template", admin: str = Depends(require_admin)):
    """엑셀 양식을 내려준다.

    mode=template  빈 양식 + 예시 3행(myquota / phd list / phd info).
    mode=current   지금 등록된 커맨드를 **같은 양식으로** 내보낸다. 엑셀에서 고쳐 그대로
                   다시 올리면 이름 기준으로 덮어써진다 - 이게 '엑셀로 수정'의 실체다.
    """
    if mode == "current":
        pool = await get_pool(_DSN)
        rows = await pool.fetch(
            "SELECT title, description, exec_command, args, host_mode, enabled, required_roles "
            "FROM execution_commands ORDER BY title")
        data = [_command_to_row(_row(r)) for r in rows]
        filename = "execution-commands.xlsx"
    else:
        data = [list(r) for r in _TEMPLATE_EXAMPLES]
        filename = "execution-commands-template.xlsx"
    content = await run_in_threadpool(_build_workbook, data)
    return Response(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ---------------------------------------------------------------- 엑셀/CSV 일괄 등록
@router.post("/commands/excel/preview")
async def preview_excel(
    file: UploadFile | None = File(None),
    server_path: str | None = Form(None),
    strip_html: bool = Form(True),
    collapse_space: bool = Form(True),
    drop_urls: bool = Form(False),
    admin: str = Depends(require_admin),
):
    ext, content, filename = await read_upload_or_server_file(file, server_path, TABLE_EXTS)
    options = {"strip_html": strip_html, "collapse_space": collapse_space, "drop_urls": drop_urls}
    upload_id = await create_upload_session(_DSN, admin, filename, ext,
                                            "execution_commands", content, options)
    session = await get_upload_session(_DSN, upload_id, admin, "execution_commands")
    try:
        sheet, header, sample, total, header_row = await run_in_threadpool(
            read_table_meta, session["saved_path"])
    except Exception as e:  # noqa: BLE001
        await delete_upload_session(_DSN, upload_id)
        raise HTTPException(422, f"파일을 읽을 수 없습니다: {e}")
    if not header:
        await delete_upload_session(_DSN, upload_id)
        raise HTTPException(422, "빈 파일입니다(헤더 행이 없습니다).")
    return {"upload_id": upload_id, "filename": filename, "sheet": sheet,
            "columns": header, "sample_rows": sample, "total_rows": total,
            "header_row": header_row, "options": options}


class ExcelCommitIn(BaseModel):
    upload_id: str
    title_column: str
    description_column: str
    exec_command_column: str | None = None


_TYPE_WORDS = {"문자열": "str", "str": "str", "string": "str", "텍스트": "str",
               "정수": "int", "int": "int", "숫자": "int", "number": "int",
               "선택형": "enum", "enum": "enum", "선택": "enum", "선택지": "enum"}
_HOST_WORDS = {"로그인 서버": "login_server", "로그인서버": "login_server",
               "login_server": "login_server", "login": "login_server",
               "대상 서버": "target_server", "대상서버": "target_server",
               "target_server": "target_server", "target": "target_server"}


def _truthy(text: str | None, default: bool) -> bool:
    t = (text or "").strip().lower()
    if not t:
        return default
    return t in ("y", "yes", "true", "1", "o", "예", "필수", "사용", "활성", "on")


def _parse_args_from_row(cell, header: list[str], placeholders: list[str]) -> list[dict]:
    """`인자N …` 열들을 인자 정의 목록으로. cell(열이름) -> 값.

    이름을 비워 두면 **실행 커맨드의 자리표시자 순서**로 채운다. 이게 이 양식의 핵심이다 -
    관리자는 `phd info {option} {job_id}`를 적고 설명만 쓰면 되고, 자리표시자 이름을
    두 번 옮겨 적지 않아도 된다(옮겨 적다 틀리면 등록이 통째로 거부된다).
    """
    slots: dict[int, dict] = {}
    for col in header:
        m = _ARG_HEADER_RE.match(col or "")
        if not m:
            continue
        slots.setdefault(int(m.group(1)), {})[m.group(2)] = (cell(col) or "").strip()

    out = []
    for pos, (idx, fields) in enumerate(sorted(slots.items())):
        name = fields.get("이름") or (placeholders[pos] if pos < len(placeholders) else "")
        if not name:
            continue
        # 이름 말고 아무것도 없고 자리표시자에도 없는 열은 빈 슬롯이다(양식의 남는 칸).
        if name not in placeholders:
            continue
        kind = _TYPE_WORDS.get((fields.get("타입") or "").strip().lower(), "str")
        choices = [c.strip() for c in _CHOICE_SPLIT.split(fields.get("선택지") or "") if c.strip()]
        if choices and kind == "str":
            kind = "enum"        # 선택지를 적었으면 선택형으로 본다(타입 칸을 안 채워도 되게)
        out.append({"name": name, "type": kind,
                    "required": _truthy(fields.get("필수"), False),
                    "default": fields.get("기본값") or "",
                    "description": fields.get("설명") or "",
                    "choices": choices})

    # 열에 안 적힌 자리표시자도 인자로 만들어 준다(설명 없이라도 있어야 등록이 통과한다).
    named = {a["name"] for a in out}
    for p in placeholders:
        if p not in named:
            out.append({"name": p, "type": "str", "required": False,
                        "default": "", "description": "", "choices": []})
    # 커맨드에 나온 순서대로 정렬한다(LLM이 보는 파라미터 순서).
    return sorted(out, key=lambda a: placeholders.index(a["name"]))


@router.post("/commands/excel/commit")
async def commit_excel(body: ExcelCommitIn, admin: str = Depends(require_admin)):
    """열 매핑으로 일괄 등록/갱신한다(title 기준 upsert).

    이름·설명·실행 커맨드는 **열을 골라서** 받는다(사내 매뉴얼에서 뽑은 표처럼 머리글이
    제각각인 파일도 올릴 수 있어야 하기 때문). 반면 인자·실행 위치·활성·역할은 양식의 머리글
    (`인자1 설명`, `실행 위치`, …)을 **이름으로 알아본다** - 고르게 할 열이 스물 몇 개가 되면
    매핑 화면이 못 쓸 물건이 된다.

    그래서 양식을 받아 쓰면 인자까지 그대로 들어오고, 머리글이 다른 예전 파일을 올리면
    예전처럼 인자 없이 들어온다(하위 호환).
    """
    session = await get_upload_session(_DSN, body.upload_id, admin, "execution_commands")
    opts = clean_options_from_dict(load_options(session))

    def _build(path: str):
        header, col_idx, rows = load_table_rows(path)
        for label, col in {"이름": body.title_column, "설명": body.description_column}.items():
            if col not in col_idx:
                raise ValueError(f"{label} 열이 파일에 없습니다: {col}")
        if body.exec_command_column and body.exec_command_column not in col_idx:
            raise ValueError(f"존재하지 않는 열입니다: {body.exec_command_column}")

        def _cell(row, col):
            if not col or col not in col_idx:
                return None
            val = row[col_idx[col]]
            return None if val is None else clean_text(str(val), opts)

        built = []
        for row in rows:
            title = _cell(row, body.title_column)
            desc = _cell(row, body.description_column)
            if not title or not desc:
                continue
            exec_command = (_cell(row, body.exec_command_column) or title).strip()
            # `{user_id}`는 예약어라 인자 표에 넣지 않는다(호출자 신원에서 자동 주입).
            placeholders = [p for p in placeholders_in(exec_command) if p != "user_id"]
            args = _parse_args_from_row(lambda c: _cell(row, c), header, placeholders)
            host_mode = _HOST_WORDS.get(
                (_cell(row, "실행 위치") or "").strip().lower(), "login_server")
            enabled = _truthy(_cell(row, "활성"), True)
            roles = [r.strip() for r in re.split(r"[,\n|]+", _cell(row, "필요 역할") or "")
                     if r.strip()]
            built.append({"title": title.strip(), "description": desc,
                          "exec_command": exec_command, "args": args,
                          "host_mode": host_mode, "enabled": enabled, "required_roles": roles})
        return built

    # 실패해도 세션을 지우지 않는다(성공 시에만 정리) - 재시도가 404로 막혀 진짜 원인이 가려진다.
    try:
        items = await run_in_threadpool(_build, session["saved_path"])
    except ValueError as e:
        raise HTTPException(422, str(e))
    await delete_upload_session(_DSN, body.upload_id)
    if not items:
        raise HTTPException(422, "등록할 커맨드가 없습니다. 이름/설명 열 선택을 확인하세요.")

    deny = await _deny()
    pool = await get_pool(_DSN)
    taken = {r["tool_name"] for r in await pool.fetch("SELECT tool_name FROM execution_commands")}
    taken |= _RESERVED_TOOL_NAMES

    inserted = updated = skipped = 0
    problems = []
    async with pool.acquire() as conn:
        async with conn.transaction():
            for it in items:
                title, exec_command = it["title"], it["exec_command"]
                existing = await conn.fetchval(
                    "SELECT tool_name FROM execution_commands WHERE title = $1", title)
                tool_name = existing or tool_name_for(title, taken, exec_command)
                try:
                    # 일괄 등록도 화면 등록과 **똑같은 검증**을 거친다. 차단 목록은 물론이고
                    # 자리표시자와 인자 정의가 맞는지도 본다(매뉴얼 표에 위험한 줄이 섞일 수
                    # 있고, 인자를 잘못 적으면 그 커맨드만 조용히 망가진다).
                    validate_definition(tool_name, exec_command, it["args"], it["host_mode"],
                                        deny, set() if existing else taken)
                except ValueError as e:
                    skipped += 1
                    if len(problems) < 10:
                        problems.append(f"{title}: {e}")
                    continue
                taken.add(tool_name)
                res = await conn.fetchrow(
                    """
                    INSERT INTO execution_commands
                        (tool_name, title, description, exec_command, args,
                         host_mode, enabled, required_roles, updated_by)
                    VALUES ($1,$2,$3,$4,$5::jsonb,$6,$7,$8,$9)
                    ON CONFLICT (title) DO UPDATE
                    SET description=EXCLUDED.description, exec_command=EXCLUDED.exec_command,
                        args=EXCLUDED.args, host_mode=EXCLUDED.host_mode,
                        enabled=EXCLUDED.enabled, required_roles=EXCLUDED.required_roles,
                        updated_by=EXCLUDED.updated_by, updated_at=now()
                    RETURNING (xmax = 0) AS inserted
                    """, tool_name, title, it["description"], exec_command,
                    json.dumps(it["args"], ensure_ascii=False),
                    it["host_mode"], it["enabled"], it["required_roles"], admin)
                if res["inserted"]:
                    inserted += 1
                else:
                    updated += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped,
            "total": len(items), "problems": problems, "restart_required": True}
