"""
VOC(사용자/운영자 질의응답 이력) 관리 API.
개별 등록/수정/삭제와, 엑셀·CSV·TSV 열 매핑 일괄 업로드를 지원한다.
질문/답변만 다룬다 - 부서·해결여부는 실제로 쓰이지 않아 화면과 API에서 제외했다
(DB 컬럼은 기존 데이터를 위해 남겨 두고 새 등록 시 기본값이 들어간다).
"""
import asyncio
import tempfile
import uuid
from datetime import datetime, timezone

import openpyxl
from fastapi import APIRouter, Depends, Form, UploadFile, File, HTTPException
from fastapi.concurrency import run_in_threadpool
from pydantic import BaseModel

from auth import require_admin
from cleaning import clean_text
from db import get_pool, embed_text, embed_texts, vector_literal
from server_files import read_upload_or_server_file
from spreadsheet import TABLE_EXTS, read_table_meta, load_table_rows
from uploads import (
    create_upload_session, get_upload_session, delete_upload_session,
)

router = APIRouter(prefix="/api/voc", tags=["voc"])

# 사내 VOC 엑셀 표준 포맷: 4행이 헤더, 이 4개 컬럼만 쓴다(의뢰번호/클러스터/의뢰자 등은 안 씀).
_VOC_HEADER_ROW = 4
_COL_REQUEST = "의뢰내용"
_COL_ACTION_DATE = "조치일"
_COL_RESOLUTION = "처리내용"
_COL_SATISFACTION = "만족도"
# 이 값이면 제외(불만족류). 매우만족/만족/보통/빈값은 그대로 사용.
_EXCLUDED_SATISFACTION = {"불만족", "매우불만족"}


class VocIn(BaseModel):
    # 부서/해결 여부는 실제로 쓰이지 않아 화면과 API에서 뺐다(DB 컬럼은 기존 데이터를 위해 유지).
    question: str
    answer: str


@router.get("")
async def list_voc(q: str | None = None, batch_id: str | None = None,
                   offset: int = 0, limit: int = 100,
                   admin: str = Depends(require_admin)):
    """VOC 목록. 수천 건이 올라가므로 반드시 페이지 단위로 준다.
    batch_id를 주면 그 업로드 묶음의 행만 본다."""
    pool = await get_pool("voc_db_dsn")
    offset = max(0, int(offset))
    limit = max(1, min(int(limit), 500))
    total = await pool.fetchval(
        """
        SELECT count(*) FROM voc_records
        WHERE ($1::text IS NULL OR question ILIKE '%' || $1 || '%')
          AND ($2::text IS NULL OR batch_id = $2)
        """,
        q or None, batch_id or None,
    )
    rows = await pool.fetch(
        """
        SELECT id, question, answer, created_at, batch_id, source_file,
               handled_by, handled_by_reason, handled_by_source
        FROM voc_records
        WHERE ($1::text IS NULL OR question ILIKE '%' || $1 || '%')
          AND ($2::text IS NULL OR batch_id = $2)
        ORDER BY created_at DESC, id DESC
        OFFSET $3 LIMIT $4
        """,
        q or None, batch_id or None, offset, limit,
    )
    return {"total": total or 0, "offset": offset, "limit": limit,
            "items": [dict(r) for r in rows]}


@router.get("/batches")
async def list_voc_batches(admin: str = Depends(require_admin)):
    """업로드(파일) 단위 묶음 목록. CSV 하나가 여기서는 한 줄로 보인다."""
    pool = await get_pool("voc_db_dsn")
    rows = await pool.fetch(
        """
        SELECT batch_id, max(source_file) AS source_file, max(uploaded_by) AS uploaded_by,
               count(*) AS record_count, min(created_at) AS uploaded_at
        FROM voc_records
        WHERE batch_id IS NOT NULL
        GROUP BY batch_id
        ORDER BY min(created_at) DESC
        """
    )
    orphan = await pool.fetchval(
        "SELECT count(*) FROM voc_records WHERE batch_id IS NULL")
    return {"batches": [dict(r) for r in rows], "unbatched": orphan or 0}


@router.delete("/batches/{batch_id}")
async def delete_voc_batch(batch_id: str, admin: str = Depends(require_admin)):
    """업로드 묶음을 통째로 삭제한다(잘못 올린 파일 되돌리기)."""
    pool = await get_pool("voc_db_dsn")
    result = await pool.execute("DELETE FROM voc_records WHERE batch_id = $1", batch_id)
    deleted = int(result.rsplit(" ", 1)[-1]) if result else 0
    return {"ok": True, "deleted": deleted}


class VocBulkDeleteIn(BaseModel):
    ids: list[int]


@router.post("/bulk-delete")
async def bulk_delete_voc(body: VocBulkDeleteIn, admin: str = Depends(require_admin)):
    """체크한 행들을 한 번에 삭제한다."""
    ids = [int(i) for i in (body.ids or [])]
    if not ids:
        raise HTTPException(422, "삭제할 항목을 선택하세요.")
    pool = await get_pool("voc_db_dsn")
    result = await pool.execute("DELETE FROM voc_records WHERE id = ANY($1::int[])", ids)
    deleted = int(result.rsplit(" ", 1)[-1]) if result else 0
    return {"ok": True, "deleted": deleted}


@router.post("")
async def create_voc(body: VocIn, admin: str = Depends(require_admin)):
    vec = await embed_text(f"{body.question}\n{body.answer}")
    pool = await get_pool("voc_db_dsn")
    row_id = await pool.fetchval(
        """
        INSERT INTO voc_records (question, answer, embedding)
        VALUES ($1, $2, $3::vector) RETURNING id
        """,
        body.question,
        body.answer,
        vector_literal(vec),
    )
    return {"id": row_id}


@router.patch("/{voc_id}")
async def update_voc(voc_id: int, body: VocIn, admin: str = Depends(require_admin)):
    vec = await embed_text(f"{body.question}\n{body.answer}")
    pool = await get_pool("voc_db_dsn")
    row = await pool.fetchrow(
        """
        UPDATE voc_records SET question=$1, answer=$2, embedding=$3::vector,
               -- 답변이 바뀌면 처리 주체 판정도 무효다. 비워서 다시 분류되게 한다(#157).
               handled_by=NULL, handled_by_reason=NULL,
               handled_by_source=NULL, handled_by_at=NULL
        WHERE id=$4 RETURNING id
        """,
        body.question,
        body.answer,
        vector_literal(vec),
        voc_id,
    )
    if not row:
        raise HTTPException(404, "VOC 기록을 찾을 수 없습니다.")
    return {"ok": True}


@router.delete("/{voc_id}")
async def delete_voc(voc_id: int, admin: str = Depends(require_admin)):
    pool = await get_pool("voc_db_dsn")
    await pool.execute("DELETE FROM voc_records WHERE id = $1", voc_id)
    return {"ok": True}


# --- 처리 주체 분류 (#157) -----------------------------------------------------------
# 수천 건을 LLM으로 훑는 작업이라 요청 안에서 끝낼 수 없다(브라우저가 먼저 끊는다).
# 백그라운드 태스크로 돌리고 진행 상황만 폴링하게 한다.
_classify_state: dict = {"running": False, "done": 0, "total": 0,
                         "classified": 0, "failed": 0, "error": "", "finished_at": None,
                         "stopping": False}
_classify_task = None


@router.get("/classify/status")
async def classify_status(admin: str = Depends(require_admin)):
    """미분류 건수와 진행 상황. 화면이 이걸 폴링한다."""
    pool = await get_pool("voc_db_dsn")
    row = await pool.fetchrow(
        """
        SELECT count(*) AS total,
               count(*) FILTER (WHERE handled_by IS NULL) AS pending,
               count(*) FILTER (WHERE handled_by = 'user') AS c_user,
               count(*) FILTER (WHERE handled_by = 'operator') AS c_operator,
               count(*) FILTER (WHERE handled_by = 'unknown') AS c_unknown
        FROM voc_records
        """
    )
    return {**dict(row), "progress": dict(_classify_state)}


@router.post("/classify")
async def classify_voc(limit: int = 0, admin: str = Depends(require_admin)):
    """미분류 VOC의 처리 주체를 LLM으로 판정해 저장한다(백그라운드).

    `limit`을 주면 그만큼만 — 수천 건을 돌리기 전에 몇 건으로 품질을 먼저 보기 위한 것이다.
    결과는 `handled_by_reason`에 남으므로 콘솔 목록에서 판정 근거를 확인할 수 있다.
    """
    global _classify_task
    if _classify_state["running"]:
        raise HTTPException(409, "이미 분류가 진행 중입니다.")

    from voc_classify import classify_pending

    pool = await get_pool("voc_db_dsn")
    _classify_state.update(running=True, done=0, total=0, classified=0, failed=0,
                           error="", finished_at=None, stopping=False)

    def progress(done, total, classified, failed):
        _classify_state.update(done=done, total=total,
                               classified=classified, failed=failed)

    async def _run():
        try:
            await classify_pending(pool, limit=limit, progress=progress,
                                   should_stop=lambda: _classify_state["stopping"])
        except Exception as e:  # noqa: BLE001
            _classify_state["error"] = f"{type(e).__name__}: {e}"
            print(f"[voc] 분류 실패: {e}")
        finally:
            _classify_state["running"] = False
            _classify_state["stopping"] = False
            _classify_state["finished_at"] = datetime.now(timezone.utc).isoformat()

    _classify_task = asyncio.create_task(_run())
    return {"ok": True, "started": True}


@router.post("/classify/stop")
async def classify_stop(admin: str = Depends(require_admin)):
    """분류를 중지한다 — **이미 나가 있는 배치는 끝내고 저장한 뒤** 멈춘다 (#161).

    태스크를 `cancel()`하지 않는 이유: LLM에 나가 있는 요청을 중간에 끊으면 그 배치의 판정은
    그냥 버려진다(돈과 시간만 쓰고 아무것도 안 남는다). 대기 중인 배치만 접으면 사용자는
    **한 건도 잃지 않고** 멈출 수 있다. 그래서 누르고 나서 몇 초 더 걸릴 수 있다.

    멈춘 뒤 남은 건은 미분류로 남아 있으므로, 다시 시작하면 그 지점부터 이어서 간다.
    """
    if not _classify_state["running"]:
        return {"ok": True, "stopping": False, "message": "진행 중인 분류가 없습니다."}
    _classify_state["stopping"] = True
    return {"ok": True, "stopping": True}


def _header_row(ws, row_num: int) -> list[str]:
    row = next(ws.iter_rows(min_row=row_num, max_row=row_num))
    return [str(c.value).strip() if c.value else "" for c in row]


# "임베딩을 안 넘긴 것"과 "임베딩을 만들지 못해 None으로 넘긴 것"을 구분하기 위한 표식.
# None을 '미지정'으로 보면, 임베딩에 실패한 행에서 여기서 또 호출해 같은 예외가 난다.
_NO_EMBEDDING = object()


async def _insert_voc(pool, question: str, answer: str, department: str | None, resolved: bool,
                      batch: dict | None = None, embedding=_NO_EMBEDDING) -> bool:
    """VOC 한 건을 저장한다. batch가 주어지면 어느 업로드에서 왔는지도 함께 남긴다
    (묶음 단위 조회·삭제용). embedding을 넘기면 그대로 쓰고(None이면 임베딩 없이 저장),
    아예 안 넘기면 여기서 만든다."""
    if not question or not answer:
        return False
    vec = (await embed_text(f"{question}\n{answer}")
           if embedding is _NO_EMBEDDING else embedding)
    b = batch or {}
    await pool.execute(
        """
        INSERT INTO voc_records (question, answer, department, resolved, embedding,
                                 batch_id, source_file, uploaded_by)
        VALUES ($1, $2, $3, $4, $5::vector, $6, $7, $8)
        """,
        question, answer, department, resolved, vector_literal(vec) if vec else None,
        b.get("batch_id"), b.get("source_file"), b.get("uploaded_by"),
    )
    return True


async def _insert_many(pool, items, batch: dict) -> tuple[int, int, list[str]]:
    """일괄 등록((question, answer, department, resolved) 튜플들). 한 건이 실패해도
    전체를 버리지 않는다.

    특정 행에서만 임베딩이 실패할 수 있다(예: 본문이 유난히 긴 건). 예전에는 그 한 건에서
    예외가 위로 튀어 **나머지 수백 건이 통째로 날아갔다**(333건 중 16건만 등록되고 500).
    지금은 실패한 행을 임베딩 없이 저장하고 계속 진행한다 - 벡터 검색에서만 빠지고
    키워드·3gram 검색에는 잡히므로 데이터를 잃지 않는다.
    임베딩 서버 자체가 죽은 경우는 embed_texts가 연속 실패를 감지해 예외를 던진다.
    """
    # 임베딩은 **묶어서** 한 번에 처리한다. 행마다 호출하면 2천 행 = 요청 2천 번이라
    # 몇 분씩 걸리고, 그동안 화면이 멈춘 것처럼 보인다.
    try:
        vectors = await embed_texts([f"{q}\n{a}" for q, a, _d, _r in items])
    except Exception as e:  # noqa: BLE001
        # 연속 실패 = 임베딩 서버 자체 문제. 아무것도 넣지 않고 중단한다.
        raise HTTPException(
            503,
            f"임베딩 서버 오류로 등록을 중단했습니다(저장된 행 없음). "
            f"임베딩 서버를 확인한 뒤 다시 등록하세요. 원인: {e}")

    inserted, skipped, failures = 0, 0, []
    for i, ((q, a, dept, resolved), vec) in enumerate(zip(items, vectors), start=1):
        if vec is None:
            # 그 행만 임베딩 없이 저장한다 - 벡터 검색에서만 빠지고 키워드·3gram에는 잡힌다.
            failures.append(f"{i}행: 임베딩 실패(본문이 너무 길거나 형식이 특이한 행)")
        if await _insert_voc(pool, q, a, dept, resolved, batch, embedding=vec):
            inserted += 1
        else:
            skipped += 1
    return inserted, skipped, failures


async def _import_raw_format(ws, header: list[str], pool, batch: dict) -> tuple[int, int]:
    """사내 VOC 표준 엑셀: 4행 헤더, 의뢰내용/조치일/처리내용/만족도만 사용."""
    col_idx = {name: i for i, name in enumerate(header)}

    def cell(row, name):
        i = col_idx[name]
        return row[i] if i < len(row) else None

    items, skipped = [], 0
    for row in ws.iter_rows(min_row=_VOC_HEADER_ROW + 1, values_only=True):
        request_content = cell(row, _COL_REQUEST)
        action_date = cell(row, _COL_ACTION_DATE)
        resolution = cell(row, _COL_RESOLUTION)
        satisfaction = cell(row, _COL_SATISFACTION)

        if not request_content or not action_date or not resolution:
            skipped += 1
            continue
        if str(satisfaction).strip() in _EXCLUDED_SATISFACTION:
            skipped += 1
            continue

        items.append((clean_text(str(request_content)), clean_text(str(resolution)), None, True))

    ins, more_skipped, _failures = await _insert_many(pool, items, batch)
    return ins, skipped + more_skipped


async def _import_simple_format(ws, header: list[str], pool, batch: dict) -> tuple[int, int]:
    """1행 헤더 Question/Answer(대소문자 무관) + department/resolved(선택), 2행부터 데이터.
    이미 정제된 텍스트로 취급하되 혹시 남은 HTML은 안전하게 걷어낸다."""
    lower_idx = {h.lower(): i for i, h in enumerate(header)}
    q_idx, a_idx = lower_idx["question"], lower_idx["answer"]
    dept_idx = lower_idx.get("department")
    resolved_idx = lower_idx.get("resolved")

    def cell(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    items, skipped = [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        question_raw, answer_raw = cell(row, q_idx), cell(row, a_idx)
        if not question_raw or not answer_raw:
            skipped += 1
            continue
        department = str(cell(row, dept_idx)) if cell(row, dept_idx) else None
        resolved_val = cell(row, resolved_idx)
        resolved = (
            str(resolved_val).strip().upper() not in ("FALSE", "0", "N", "NO")
            if resolved_val is not None else True
        )

        items.append((clean_text(str(question_raw)), clean_text(str(answer_raw)),
                      department, resolved))

    ins, more_skipped, _failures = await _insert_many(pool, items, batch)
    return ins, skipped + more_skipped


@router.post("/import")
async def import_voc_excel(
    file: UploadFile | None = File(None),
    server_path: str | None = Form(None),
    admin: str = Depends(require_admin),
):
    """엑셀 형식을 자동으로 인식해서 등록한다. 지원하는 두 형식:
    (1) 1행 헤더 Question/Answer(대소문자 무관, department/resolved 선택) — 이미 정제된 데이터용.
    (2) 사내 VOC 표준 포맷 — 4행 헤더(의뢰내용/조치일/처리내용/만족도), 조치일·처리내용 있는 행만,
        만족도 불만족/매우불만족 제외, 본문은 HTML 태그만 벗기고 그대로 보존."""
    _, content, filename = await read_upload_or_server_file(file, server_path, {".xlsx"})
    # 이 업로드에서 들어간 행들을 하나로 묶는다(콘솔에서 파일 단위로 보고 되돌리기 위함).
    batch = {"batch_id": uuid.uuid4().hex, "source_file": filename, "uploaded_by": admin}
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    wb = openpyxl.load_workbook(tmp_path, read_only=True)
    ws = wb.active
    pool = await get_pool("voc_db_dsn")

    row1 = _header_row(ws, 1)
    if {"question", "answer"}.issubset({h.lower() for h in row1}):
        inserted, skipped = await _import_simple_format(ws, row1, pool, batch)
    elif ws.max_row >= _VOC_HEADER_ROW and {
        _COL_REQUEST, _COL_ACTION_DATE, _COL_RESOLUTION, _COL_SATISFACTION
    }.issubset(set(_header_row(ws, _VOC_HEADER_ROW))):
        inserted, skipped = await _import_raw_format(
            ws, _header_row(ws, _VOC_HEADER_ROW), pool, batch)
    else:
        raise HTTPException(
            422,
            "엑셀 형식을 인식하지 못했습니다. 지원 형식: "
            "(1) 1행 헤더 Question/Answer, 2행부터 데이터, 또는 "
            f"(2) {_VOC_HEADER_ROW}행 헤더에 {_COL_REQUEST}/{_COL_ACTION_DATE}/"
            f"{_COL_RESOLUTION}/{_COL_SATISFACTION} 컬럼이 모두 있는 사내 표준 포맷.",
        )

    return {"inserted": inserted, "skipped": skipped, "batch_id": batch["batch_id"]}


# ---------------------------------------------------------------- 열 매핑 업로드(형식 자유)
# 고정된 두 포맷(1행 Question/Answer · 4행 사내표준)만 받던 것을 대체한다.
# 헤더 행을 자동으로 찾고(제목 줄이 위에 몇 개 있어도 됨), 어떤 열을 무엇으로 쓸지 고르게 한다.
_DSN_VOC = "voc_db_dsn"


def _guess(columns: list[str], candidates: list[str]) -> str:
    for c in columns:
        low = c.lower()
        if any(k in low for k in candidates):
            return c
    return ""


@router.post("/excel/preview")
async def preview_voc_table(
    file: UploadFile | None = File(None),
    server_path: str | None = Form(None),
    header_row: int | None = Form(None),
    admin: str = Depends(require_admin),
):
    """엑셀/CSV/TSV의 헤더 행을 자동으로 찾아 열 목록과 샘플을 돌려준다.
    header_row를 주면 그 행(1-based, 엑셀에서 보이는 실제 행 번호)을 헤더로 강제한다."""
    ext, content, filename = await read_upload_or_server_file(file, server_path, TABLE_EXTS)
    upload_id = await create_upload_session(_DSN_VOC, admin, filename, ext, "voc_table", content, {})
    session = await get_upload_session(_DSN_VOC, upload_id, admin, "voc_table")
    try:
        sheet, header, sample, total, detected = await run_in_threadpool(
            read_table_meta, session["saved_path"], 5, header_row)
    except Exception as e:  # noqa: BLE001
        await delete_upload_session(_DSN_VOC, upload_id)
        raise HTTPException(422, f"파일을 읽을 수 없습니다: {e}")
    if not header:
        await delete_upload_session(_DSN_VOC, upload_id)
        raise HTTPException(422, "빈 파일입니다(표를 찾지 못했습니다).")

    return {
        "upload_id": upload_id, "filename": filename, "sheet": sheet,
        "columns": header, "sample_rows": sample, "total_rows": total,
        "header_row": detected,
        # 사내 표준 포맷이면 매핑을 미리 채워준다(그대로 등록만 누르면 되게).
        "suggest": {
            "question_column": _guess(header, ["의뢰내용", "question", "문의", "질문", "요청"]),
            "answer_column": _guess(header, ["처리내용", "answer", "답변", "조치", "회신"]),
        },
    }


class VocTableCommitIn(BaseModel):
    """require_columns: 이 열들이 비어 있는 행은 건너뛴다(예: 조치일이 없는 미처리 건)."""
    upload_id: str
    header_row: int | None = None
    question_column: str
    answer_column: str
    require_columns: list[str] = []


@router.post("/excel/commit")
async def commit_voc_table(body: VocTableCommitIn, admin: str = Depends(require_admin)):
    session = await get_upload_session(_DSN_VOC, body.upload_id, admin, "voc_table")

    def _build(path: str):
        header, col_idx, rows = load_table_rows(path, body.header_row)
        for label, col in (("질문", body.question_column), ("답변", body.answer_column)):
            if col not in col_idx:
                raise ValueError(f"{label} 열이 파일에 없습니다: {col}")
        for col in body.require_columns:
            if col and col not in col_idx:
                raise ValueError(f"존재하지 않는 열입니다: {col}")

        def cell(row, col):
            if not col or col not in col_idx:
                return None
            v = row[col_idx[col]]
            return None if v is None else str(v).strip()

        built, skipped = [], 0
        for row in rows:
            if any(not cell(row, c) for c in body.require_columns):
                skipped += 1
                continue
            q, a = cell(row, body.question_column), cell(row, body.answer_column)
            if not q or not a:
                skipped += 1
                continue
            built.append((clean_text(q), clean_text(a), None, True))
        return built, skipped

    # 실패해도 세션을 지우지 않는다. 예전에는 finally로 무조건 지워서, 등록이 한 번
    # 실패하면(열 선택 오류·임베딩 서버 오류 등) 재시도할 때 "업로드 세션이 없거나
    # 만료되었습니다"(404)가 떠 **진짜 원인이 404로 덮였다.** 성공했을 때만 정리하고,
    # 실패한 세션은 TTL(기본 60분)이 알아서 치운다.
    try:
        items, skipped = await run_in_threadpool(_build, session["saved_path"])
    except ValueError as e:
        raise HTTPException(422, str(e))

    if not items:
        raise HTTPException(422, "등록할 행이 없습니다. 열 선택과 제외 조건을 확인하세요.")

    pool = await get_pool(_DSN_VOC)
    # 이 업로드에서 들어간 행들을 하나로 묶는다(콘솔에서 파일 단위로 보고 되돌리기 위함).
    batch = {"batch_id": uuid.uuid4().hex,
             "source_file": session["filename"], "uploaded_by": admin}
    inserted, more_skipped, failures = await _insert_many(pool, items, batch)
    skipped += more_skipped

    await delete_upload_session(_DSN_VOC, body.upload_id)
    return {"inserted": inserted, "skipped": skipped, "total": len(items),
            "embed_failed": len(failures), "embed_errors": failures[:10],
            "batch_id": batch["batch_id"], "source_file": session["filename"]}
