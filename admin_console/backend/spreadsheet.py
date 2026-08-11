"""
표 형식 파일(.xlsx/.xls/.csv/.tsv) 읽기 공통 유틸.

여러 탭(매뉴얼·커맨드 등)이 같은 방식으로 헤더/샘플을 미리 보고, 선택한 열을 정제해
가져올 수 있도록 파싱을 한 곳에 모은다. 실제 정제(clean_text)는 호출부에서 수행한다.

CSV/TSV는 엑셀과 완전히 같은 "열 매핑" 흐름을 타도록 여기서 흡수한다(호출부는 확장자를 몰라도 된다).
사내에서 만든 CSV는 UTF-8(BOM 포함)이거나 CP949(엑셀 한글 기본)인 경우가 많아 둘 다 시도하고,
구분자도 쉼표/탭/세미콜론/파이프를 자동 판별한다. 단 **.tsv는 탭으로 고정**한다 -
셀 안에 쉼표가 흔해서 자동 판별에 맡기면 쉼표를 구분자로 오인하는 경우가 있다.
"""
import csv
import os

import openpyxl

CSV_EXTS = {".csv"}
TSV_EXTS = {".tsv"}
TEXT_TABLE_EXTS = CSV_EXTS | TSV_EXTS
EXCEL_EXTS = {".xlsx", ".xls"}
TABLE_EXTS = EXCEL_EXTS | TEXT_TABLE_EXTS

_CSV_ENCODINGS = ("utf-8-sig", "cp949", "utf-8", "latin-1")

# csv 모듈의 기본 필드 상한은 128KB라, 셀 하나가 그보다 크면
# `field larger than field limit (131072)`로 파일 전체를 읽지 못한다.
# VOC 처리내용처럼 긴 본문 한 칸이 이 한도를 넘는 경우가 실제로 있다.
# 업로드 자체가 upload_max_mb(기본 50MB)로 제한되므로, 그보다 넉넉한 값으로 올려 둔다
# (플랫폼에 따라 C long을 넘으면 OverflowError가 나므로 절반씩 낮추며 시도).
def _raise_field_size_limit(target: int = 64 * 1024 * 1024) -> int:
    limit = target
    while limit > 128 * 1024:
        try:
            csv.field_size_limit(limit)
            return limit
        except (OverflowError, ValueError):
            limit //= 2
    return csv.field_size_limit()


CSV_FIELD_SIZE_LIMIT = _raise_field_size_limit()


def _read_csv_all(path: str) -> list[list[str]]:
    """CSV/TSV 전체를 문자열 2차원 리스트로 읽는다. 인코딩/구분자를 자동으로 맞춘다."""
    is_tsv = os.path.splitext(path)[1].lower() in TSV_EXTS
    last_err: Exception | None = None
    for enc in _CSV_ENCODINGS:
        try:
            with open(path, "r", newline="", encoding=enc) as f:
                if is_tsv:
                    # 확장자가 탭 구분을 명시하므로 판별하지 않는다. VOC 본문에는 쉼표가 흔해서
                    # 자동 판별에 맡기면 쉼표를 구분자로 잡아 열이 어긋난다.
                    return [row for row in csv.reader(f, csv.excel_tab)]
                sample = f.read(8192)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel        # 한 열짜리 파일 등 판별 실패 시 기본(쉼표)
                return [row for row in csv.reader(f, dialect)]
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except csv.Error as e:
            # 인코딩 문제가 아니라 형식 문제이므로 다른 인코딩으로 재시도해도 소용없다.
            raise _friendly_csv_error(e)
    raise ValueError(
        f"파일 인코딩을 인식할 수 없습니다(UTF-8 또는 CP949로 저장해 주세요): {last_err}")


def _friendly_csv_error(e: csv.Error) -> ValueError:
    msg = str(e)
    if "field larger than field limit" in msg:
        return ValueError(
            f"한 칸의 내용이 너무 커서 읽지 못했습니다(현재 상한 {CSV_FIELD_SIZE_LIMIT:,}자). "
            "따옴표가 짝이 맞지 않아 여러 행이 한 칸으로 붙었을 가능성이 큽니다 - "
            f"원본 파일의 따옴표(\")를 확인해 주세요. 원인: {msg}")
    return ValueError(f"파일을 파싱하지 못했습니다: {msg}")


def _is_csv(path: str) -> bool:
    """구분자 텍스트 표(csv/tsv)인지. 엑셀이면 False."""
    return os.path.splitext(path)[1].lower() in TEXT_TABLE_EXTS


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def detect_header_row(rows: list[list], max_scan: int = 15) -> int:
    """헤더로 보이는 행의 인덱스(0-based)를 추정한다.

    엑셀마다 1행부터 표가 시작하기도 하고, 2행에 제목이 있고 4행부터 표가 나오기도 한다.
    "헤더 행은 (1) 채워진 칸이 여러 개고 (2) 각 칸이 짧은 라벨이며 (3) 값이 서로 다르고
    (4) 바로 아래에 비슷한 폭의 데이터 행이 이어진다"는 성질로 점수를 매겨 가장 높은 행을 고른다.
    """
    best_idx, best_score = 0, -1.0
    for i, row in enumerate(rows[:max_scan]):
        cells = [_norm(v) for v in row]
        filled = [c for c in cells if c]
        if len(filled) < 2:
            continue
        # 제목 줄(한 칸에 긴 문장)이 헤더로 뽑히지 않도록 긴 칸에 벌점
        long_cells = sum(1 for c in filled if len(c) > 40)
        distinct = len(set(filled)) / len(filled)
        # 아래로 3행이 이 폭을 유지하는지(데이터가 실제로 이어지는지)
        follow = 0
        for r in rows[i + 1:i + 4]:
            if sum(1 for v in r if _norm(v)) >= max(2, len(filled) * 0.6):
                follow += 1
        score = len(filled) + distinct * 2 + follow * 1.5 - long_cells * 3
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _header_of(row) -> list[str]:
    return [str(v).strip() if v is not None and str(v).strip() else f"column_{i}"
            for i, v in enumerate(row)]


def _all_rows(path: str) -> list[list]:
    """엑셀/CSV/TSV를 2차원 리스트로 읽는다. 빈 행도 그대로 둔다 —
    그래야 헤더 행 번호가 사용자가 엑셀에서 보는 실제 행 번호와 일치한다."""
    if _is_csv(path):
        return [list(r) for r in _read_csv_all(path)]
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return [list(r) for r in wb.active.iter_rows(values_only=True)]
    finally:
        wb.close()


def _sheet_name(path: str) -> str:
    if _is_csv(path):
        return os.path.splitext(path)[1].lower().lstrip(".").upper()
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return wb.active.title
    finally:
        wb.close()


def read_table_meta(path: str, sample_size: int = 5, header_row: int | None = None):
    """(sheet, header, sample_rows, total_rows, header_row)를 반환한다.

    header_row(1-based)를 주지 않으면 자동으로 찾는다 — 1행부터 표가 시작하는 파일도,
    위에 제목/설명 줄이 몇 개 있고 중간부터 표가 시작하는 파일도 그대로 받기 위함이다.
    """
    rows = _all_rows(path)
    if not rows:
        return None, [], [], 0, 1
    idx = (header_row - 1) if header_row else detect_header_row(rows)
    idx = max(0, min(idx, len(rows) - 1))
    header = _header_of(rows[idx])
    body = [r for r in rows[idx + 1:] if any(_norm(v) for v in r)]
    sample = [[_norm(v) for v in r] for r in body[:sample_size]]
    return _sheet_name(path), header, sample, len(body), idx + 1


def load_table_rows(path: str, header_row: int | None = None):
    """(header, col_idx, rows)를 반환한다. col_idx는 열 이름 -> 인덱스 매핑이다.
    header_row(1-based)를 주지 않으면 read_table_meta와 같은 방식으로 자동 판별한다."""
    rows = _all_rows(path)
    if not rows:
        return [], {}, []
    idx = (header_row - 1) if header_row else detect_header_row(rows)
    idx = max(0, min(idx, len(rows) - 1))
    header = _header_of(rows[idx])
    col_idx = {name: i for i, name in enumerate(header)}
    # 헤더보다 열이 적은 행이 있어도 인덱스 접근이 터지지 않게 길이를 맞춘다.
    body = [r + [None] * (len(header) - len(r)) if len(r) < len(header) else r
            for r in rows[idx + 1:] if any(_norm(v) for v in r)]
    return header, col_idx, body


# 이전 이름 유지(호출부 점진 이행용).
read_excel_meta = read_table_meta
load_excel_rows = load_table_rows
